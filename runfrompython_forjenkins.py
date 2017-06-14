from subprocess import call
import os
from GlassDemoDataProvider import GlassDemoDataProvider
import datetime
import time
from shutil import copyfile
import openpyxl
import xlrd

class RunRobotFromPython():
    executionMode="jenkins"#Debugg|batch_run|jenkins
    directory=''
    Date=""
    testCasefileName=''
    reservedBoxes=''
    guideVerificationFailed=''
    screenShotprefix=''
    statictext1=''
    statictext2=''
    SlotIDPublic=''
    reportPath=""

##    
##    scroll_down=(str(GlassDemoDataProvider().getCoordinates('scroll_down',"x")),str(GlassDemoDataProvider().getCoordinates('scroll_down',"y")))
##    standby=(str(GlassDemoDataProvider().getCoordinates('standby',"x")),str(GlassDemoDataProvider().getCoordinates('standby',"y")))
##    open_fullscreen=(str(GlassDemoDataProvider().getCoordinates('open_fullscreen',"x")),str(GlassDemoDataProvider().getCoordinates('open_fullscreen',"y")))
##    exit_fullscreen=(str(GlassDemoDataProvider().getCoordinates('exit_fullscreen',"x")),str(GlassDemoDataProvider().getCoordinates('exit_fullscreen',"y")))
##    Guide=(str(GlassDemoDataProvider().getCoordinates('Guide',"x")),str(GlassDemoDataProvider().getCoordinates('Guide',"y")))
##    exitkey=(str(GlassDemoDataProvider().getCoordinates('exitkey',"x")),str(GlassDemoDataProvider().getCoordinates('exitkey',"y")))
##    down=(str(GlassDemoDataProvider().getCoordinates('down',"x")),str(GlassDemoDataProvider().getCoordinates('down',"y")))
##    menu=(str(GlassDemoDataProvider().getCoordinates('menu',"x")),str(GlassDemoDataProvider().getCoordinates('menu',"y")))
##    up=(str(GlassDemoDataProvider().getCoordinates('up',"x")),str(GlassDemoDataProvider().getCoordinates('up',"y")))
##    select=(str(GlassDemoDataProvider().getCoordinates('select',"x")),str(GlassDemoDataProvider().getCoordinates('select',"y")))
##    right=(str(GlassDemoDataProvider().getCoordinates('right',"x")),str(GlassDemoDataProvider().getCoordinates('right',"y")))
##    left=(str(GlassDemoDataProvider().getCoordinates('left',"x")),str(GlassDemoDataProvider().getCoordinates('left',"y")))
##    scroll_up=(str(GlassDemoDataProvider().getCoordinates('scroll_up',"x")),str(GlassDemoDataProvider().getCoordinates('scroll_up',"y")))
##    key_one=(str(GlassDemoDataProvider().getCoordinates('key_one',"x")),str(GlassDemoDataProvider().getCoordinates('key_one',"y")))
##    key_two=(str(GlassDemoDataProvider().getCoordinates('key_two',"x")),str(GlassDemoDataProvider().getCoordinates('key_two',"y")))
##







    
    def createTimestamp(self):
        """this function will make the timestamp of execution and which will be used to create the execution folder. No input required """
        ts = time.time()
        self.directory = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H-%M-%S')
        self.Date= datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
        print self.Date
        print self.directory 
    def ceateExecutionDirectory(self):
        """This function is making the execution directory.No input required """
        if not os.path.exists(self.directory):
            os.makedirs(self.directory)
            print "execution directory created"
        else:
            print 'Error in making execution directory'
    def createScreenshotDirectory(self):
        if os.path.exists(self.directory):
               os.makedirs(self.directory+'//screenshot')
               print "screenshot directory created"
        else: print 'screen shot folder not created'
       
   

    def createTestCaseFile(self):
        """ This function is creating Testcase file for writing test cases dynamiccaly. No input required """
        self.testCasefileName=self.directory+'/Runfrompy.txt'
        print self.testCasefileName," is located" 
    def createReservedBoxes(self):
        """Ceating the file to make the list of reserved boxes"""
        self.reservedBoxes=self.directory+'/ReservedBoxes.txt'
        reservedBoxesfile = open(self.reservedBoxes,'a+')
        reservedBoxesfile.write("==========================\n")
        reservedBoxesfile.write ('Reserved boxes are below\n')
        reservedBoxesfile.write("==========================\n")
        reservedBoxesfile.close
        #<metigation steps for avoiding delay to write to file>
        reservedBoxesfile = open(self.reservedBoxes,'a+')
        reservedBoxesfile.close
        #</metigation steps for avoiding delay to write to file>
    def createListOfGuideFailedBoxes(self):
        """Ceating the file to make the list of  boxes which failed to verify Guide screen .No input required """
        self.guideVerificationFailed=self.directory+'/guideVerificationFailed.txt'
        guideVerificationFailedListFile = open(self.guideVerificationFailed,'a+')
        guideVerificationFailedListFile.write("===========================================\n")
        guideVerificationFailedListFile.write ('Guide verication failed on below boxes\n')
        guideVerificationFailedListFile.write("===========================================\n")
        guideVerificationFailedListFile.close
        #<metigation steps for avoiding delay to write to file>
        guideVerificationFailedListFile = open(self.guideVerificationFailed,'a+')
        guideVerificationFailedListFile.close
        #</metigation steps for avoiding delay to write to file>
    def createReportOverviewFile(self):
        reportOverView=open("testData/reportOverView.txt",'w+')
        reportOverView.close()
    def getScreenShotPrefix(self):
        """Extracting the screen shot prefix to get last screen shot taken. this will be used in Imagecompare Library. No input required """
        temp=self.testCasefileName.split('.')
        if '/' in temp[0]:
            temp = temp[0].split("/")
        elif '\\' in temp[0]:
            temp = temp[0].split("\\")
        self.screenShotprefix= str(temp[1])+"-screenshot-"


    def setStaticText(self):
        
        
#below are the static text using in test case file.
        self.statictext1= '''| *Setting*  |     *Value*     										|
| Library    | OperatingSystem 										|
| Library    |      Selenium2Library								|
| Library    |      OperatingSystem									|
| Library    |      Collections										|
| Library    |      ImageHorizonLibrary	| reference_folder=images	| screenshot_folder='''+self.directory+'''//screenshot |
| Library    |      imagecompare |
#below are the variables used in testcase
| *Variable* |     *Value*     |
| ${SiteUrl} | https://glas.charterlab.com/#/
| ${Username} | P2702231 |
| ${Password} | xxxxxxxxxx |
| ${Browser} | chrome |
| ${DashboardTitle}	| Lab Glas |
#below values should not be changed. Please add new variable if some new value is required
| ${DelayS}  | 1s |
| ${Delay}   | 7s |
| ${DelayM}  | 10s |
| ${DelayL}  | 20s |
| ${x} 		| 0 |
| ${count} | 0 |
| ${continueTest} | True |
| ${boxtype} | 0 |
| ${SlotID} | 0 |
#beloaw are the testcases going to execute
| *Test Case*  | *Action*        | *Argument*   |
| Login to GLAS | [Documentation] | loging to GLAS |
| 				| Open Browser to the GLAS login page |
|				| Enter User Name and password |
|				| sleep | ${Delay} |
#Eliminate Chromedriver Popup - Workaround
| 				| @{coordinates2}= | Create List | ${'''+str( GlassDemoDataProvider().getCoordinates('chromedriver_Popup',"x"))+'''} | ${'''+str( GlassDemoDataProvider().getCoordinates('chromedriver_Popup',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinates2} | 1 |
|				| set_window_size | 480 | 320 |
|				| sleep | ${Delay} |
|				| Maximize Browser Window |
|				| sleep | ${Delay} |
|				| Assert Home Screen Title |
#|				| sleep | ${DelayM} | #debugging purpos ,can be deleted
'''

        
        self.statictext2= '''
| *Keyword*     |
| Assert Home Screen Title |
|				| Page Should Contain |  Lab Glas |
| Open Browser to the GLAS login page |
|				| open browser | ${SiteUrl} | ${Browser} |
|				| Maximize Browser Window |

#|				| ${chrome_options}= |  Evaluate | sys.modules['selenium.webdriver'].ChromeOptions() |  sys |
#|				| Call Method | ${chrome_options} | add_argument | test-type |
#|				| Call Method | ${chrome_options} | add_argument | --disable-extensions |
#|				| Run Keyword If |  os.sep == '/'   |   Create Webdriver |  Chrome | my_alias | chrome_options=${chrome_options} | executable_path=/usr/local/bin/chromedriver |
#|				|  | ELSE | Create Webdriver    |   Chrome | my_alias | chrome_options=${chrome_options} | 
#|				| Go To| ${SiteUrl} |
|End Reservation |
| 				| Click Element | endReservationButton |
| 				| sleep | ${Delay} |
| 				| scroll up |
| Reserve the box |
| 				| log | Reserving box for 30 minutes |
|				| sleep | ${Delay} |
| 				| Click Element | css = #ReservationOverlay > div > div.reservation-available.reservation-panel > div > div > span:nth-child(1) |
| 				| sleep | ${Delay} |
| Signal Check 	| 
| 				| press standby |
| 				| sleep	| ${Delay} |
| 				| press standby |
| 				| sleep	| ${DelayM} |
| 				| sleep	| ${Delay} |
| 				| Take A Screenshot |
| 				| ${isSignalAvailable}= | imagecompare.verifyNOSIGNAL_UNSUPPORTEDSIGNAL | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| log | ${isSignalAvailable} |
| 				| Run Keyword If | '${isSignalAvailable}' == 'False' | Signal Check after standby |
| 				| Run Keyword If | '${isSignalAvailable}' == 'Booting' | imagecompare.ReportTestStatus | '''+self.executionMode+''' | 
| 				| Run Keyword If | '${isSignalAvailable}' == 'Booting' | Fail | msg=*HTML* ${isSignalAvailable} |
| Signal Check after standby	|
| 				| sleep	| ${DelayS} |
| 				| press standby |
| 				| sleep	| ${DelayM} |
| 				| Take A Screenshot | 
| 				| ${continueTest}= | imagecompare.verifyNOSIGNAL_UNSUPPORTEDSIGNAL | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| Run Keyword If | '${continueTest}' == 'False' | Report NO IR or NO SIGNAL or UNSUPPORTED SIGNAL |
| 				| imagecompare.setcontinueTest | ${continueTest} |
| Report NO IR or NO SIGNAL or UNSUPPORTED SIGNAL |
|				| sleep | ${DelayS} |
| 				| imagecompare.reportNOSIGNAL_UNSUPPORTEDSIGNAL_NOIR | 

| Start Verifications |
| 				| Run Keyword And Continue On Failure | Verify Guide screen |
| 				| sleep	| ${Delay} |
| 				| Run Keyword And Ignore Error |  press exit key |
| 				| sleep	| ${DelayS} |
| 				| Run Keyword And Ignore Error |  press key two |
|				| sleep | ${DelayM} |
| 				| Run Keyword And Continue On Failure | Verify ondemand screen |
| 				| sleep	| ${DelayS} |
|				| Run Keyword And Continue On Failure | verify video |
| 				| sleep	| ${DelayS} |
| 				| Run Keyword And Ignore Error | End Reservation |
| verify video 	|
| 				| press exit key  |
| 				| sleep | ${DelayS} |
| 				| press exit key  |
| 				| sleep | ${DelayS} |
| 				| press exit key  |
#| 				| scroll down |  # time bieng ,we are not verifying video in full screen <remove comment to eanble full screen verivfication>
#|				| Open in fullscreen |#<remove comment to eanble full screen verivfication>
| 				| verify screenshots |
| 				| sleep | ${Delay} |
#| 				| exit fullscreen | #<remove comment to eanble full screen verivfication>
| scroll down 	|
| 				| @{coordinatesScrollDown}= | Create List | ${'''+str( GlassDemoDataProvider().getCoordinates('scroll_down',"x"))+'''} | ${'''+str( GlassDemoDataProvider().getCoordinates('scroll_down',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesScrollDown} | 1 |
| press standby	|
| 				| @{coordinatesStandBy}= | Create List | ${'''+ str(GlassDemoDataProvider().getCoordinates('standby',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('standby',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesStandBy} | 1 |
| 				| log | power butten pressed |
| Open in fullscreen |
|				| scroll down |
| 				| @{coordinatesVideo}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('AVscreen',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('AVscreen',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesVideo} | 100 |
|				| sleep | ${DelayS} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesVideo} | 100 |
| 				| @{coordinatesFullScreen}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('open_fullscreen',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('open_fullscreen',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesFullScreen} | 1 |
|				| sleep | ${DelayM} |
| Delete Old Screen shots |
|				| imagecompare.DeleteAllOldScreenshotsin | '''+self.directory+'''/screenshot |
| Full screen Detection |
| 				| Take A Screenshot |
| 				| ${normalScreen}= | imagecompare.detectnormalScreen | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| Run Keyword If | '${normalScreen}' == 'False' | exit fullscreen  |
| verify screenshots |
|				| Take A Screenshot |
| 				| sleep | ${DelayM} |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.last2ScreenshotsCompareForAV | '''+self.screenShotprefix+''' | '''+self.directory+'''/screenshot |
| 				| log 	| ${pass}  |
#| 				| exit fullscreen |
#| 				| scroll up |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass}  |
| try again verifying screenshots after reload | 
| 				| exit fullscreen |
| 				| Reload GLAS |
| 				| Open in fullscreen |
| exit fullscreen | 
| 				| @{coordinatesExitFullScreen}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('exit_fullscreen',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('exit_fullscreen',"y"))+'''}  |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesExitFullScreen} | 1 | 
| Reload GLAS 	| 
| 				| log | reloading the Glas |  
| 				| Click Element | reloadStreamButton |
| 				| sleep | ${DelayL} | 
| Verify Guide screen |
| 				| sleep | ${DelayM}
| 				| press Guide key |
| 				| sleep	| ${DelayM} |
| 				| press key One |
| 				| sleep	| ${DelayL} |
| 				| sleep	| ${Delay} |
|				| Take A Screenshot |
|				| ${SlotID} | imagecompare.getCurrentSlotID | 
|				| Log | ${SlotID} |
| 				| ${pass}= | imagecompare.GuideVerification | Runfrompy-screenshot- | ${SlotID} | '''+self.directory+'''/screenshot |
| 				| log | ${pass} |
| 				| sleep | ${Delay} |
| 				| Run Keyword And Ignore Error | press exit key |
| 				| sleep	| ${DelayS} |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass}  |
| Verify ondemand screen |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.verifyBoxAV | Runfrompy-screenshot- | '''+self.directory+'''/screenshot 
| 				| Run Keyword If | '${pass}' == 'True' | Verify ondemand screen in fullscreen |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass}  |

| Verify ondemand screen in fullscreen |
|				| press Ondemand key |
| 				| sleep | ${DelayL} |
| 				| Open in fullscreen | 
| 				| sleep | ${DelayS} |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.onDemandVerification | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| log | ${pass} | 
| 				| exit fullscreen |
| 				| scroll up |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass}  |
| 				| ${boxtype}= | imagecompare.getBoxType | '''+self.screenShotprefix+''' | '''+self.directory+'''/screenshot |
| 				| log | ${boxtype} |
#|				| Run Keyword If | '${boxtype}' == 'spectrum' | verify OD play back  |
#|				| Run Keyword If | '${boxtype}' == 'First Time Spectrum' | verify OD play back by searching asset  |
| 				| sleep	| ${Delay} |
| 				| press exit key |
| 				| sleep	| ${DelayS} |
| 				| press exit key  |
| 				| sleep	| ${DelayS} |
| 				| press exit key  |
| 				| sleep	| ${DelayS} |
| verify search screen |
|				| Take A Screenshot |
| 				| ${search}= | imagecompare.VerifySearchScreen | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |				
| 				| Run Keyword If | '${search}' == 'False' | Fail | msg=*HTML* ${pass}  |
| verify OD play back |
| 				| press down key |
| 				| sleep | ${DelayS} |
| 				| press down key |
| 				| sleep | ${Delays} |
| 				| Take A Screenshot |
| 				| press select key  |
| 				| sleep | ${DelayM} |
| 				| press select key  |
| 				| sleep | ${Delays} |
| 				| press select key  |
| 				| sleep | ${Delays} |
| 				| sleep | ${Delay} |
| 				| press right key |
| 				| sleep | ${Delays} |
| 				| press select key  |
| 				| sleep | ${Delays} |
| 				| press select key  |
| 				| sleep | ${DelayM} |
| 				| Take A Screenshot |
|				| Take A Screenshot |
| 				| sleep | ${DelayM} |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.last2ScreenshotsCompareForAV | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| log 	| ${pass}  |
| 				| Run Keyword If | '${pass}' == 'False' | verify OD play back by searching asset  |
| report VOD Playback failure  |
| 				| imagecompare.reportVODPlaybackFailure |			



|  verify OD play back by searching asset |
| 				| press exit key  |
| 				| sleep | ${DelayS} |
| 				| press exit key  |
| 				| sleep | ${DelayS} |
| 				| press exit key  |
| 				| sleep | ${DelayS} |
| 				| press exit key  |
| 				| sleep | ${DelayS} |
| 				| search screen invocation |
| 				| sleep | ${DelayM} |
|				| verify search screen |
| 				| sleep | ${Delay} |
| 				| type empire |
|				| Take A Screenshot |
| 				| sleep | ${Delay} |
| 				| select empire for play |
| 				| sleep | ${Delay} |
| 				| press select key  |   
| 				| ${pass}= | imagecompare.last2ScreenshotsCompareForAV | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
|				| Run Keyword If | '${pass}' == 'False' | report VOD Playback failure  |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass}  |
| search screen invocation |
| 				| press menu key |
| 				| sleep | ${Delay} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
#below are the remot key configuatation and clicking the same button
| press Guide key |
| 				| @{coordinatesGuide}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('Guide',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('Guide',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesGuide} | 1 | 
|				| log | exit key pressed |
| press exit key |
| 				| @{coordinatesExit}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('exit',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('exit',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesExit} | 1 | 
|				| log | exit key pressed | 
| press Ondemand key |
| 				| @{coordinates}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('ondemand',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('ondemand',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinates} | 1 | 
|				| log | Ondemand Key pressed | 
| press down key |
| 				| @{coordinatesDown}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('down',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('down',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesDown} | 1 | 
| 				| log | down key pressed |
| press menu key |
| 				| @{coordinatesMenu}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('menu',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('menu',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesMenu} | 1 | 
| 				| log | menu key pressed |
| press up key 	|
| 				| @{coordinatesUp}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('up',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('up',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesUp} | 1 | 
| 				| log | up key pressed |
| press select key |
| 				| @{coordinatesSelect}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('select',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('select',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesSelect} | 1 | 
| 				| log | select key pressed |
| press right key |
| 				| @{coordinatesRight}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('right',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('right',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesRight} | 1 | 
| 				| log | select key pressed |
| press left key |
| 				| @{coordinatesLeft}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('left',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('left',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesLeft} | 1 | 
| 				| log | select key pressed |
| select empire for play |
| 				| sleep | ${DelayS} |
|				|  press down key  |
#| here more logic should be added to check watch option is available or not.

| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayM} |
|				| Take A Screenshot |
#browsing episodes
#verify Browse episodes for analisys. dont fail the test case
#| 				| ${browseEpisode}= | imagecompare.verifyBrowseEpisode | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| press select key  |
#| 				| sleep | ${DelayM} |
|				| Take A Screenshot |
#available to watch now
#| 				| ${AvailableToWatchNow}= | imagecompare.verifyAvailableToWatchNow | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| press select key  |
| 				| sleep | ${DelayM} |
|				| Take A Screenshot |
#synopsys screeen
| 				| ${synopsys}= | imagecompare.verifyEmpireSynopsysScreen | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |
| 				| ${watch}= | imagecompare.searchForWatch | Runfrompy-screenshot- | '''+self.directory+'''/screenshot |		
| 				| Run Keyword If | '${synopsys}' == 'False' | Fail | msg=*HTML* ${pass}  |
#| 				| Run Keyword If | '${watch}' == 'False' | Fail | msg=*HTML* ${pass}  |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| press select key  |

| type empire |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
|				|  press down key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
|				|  press down key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press left key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press left key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press left key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press left key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press left key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
|				|  press down key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press up key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press left key   |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press right key   |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
| 				| press select key  |
| 				| sleep | ${DelayS} |
| 				| sleep | ${DelayS} |
#typing empire complete
| power off the box | 
|				| Click Element | powerOffButton |
| 				| sleep | ${DelayL} |
| 				| ImageHorizonLibrary.Does Exist | poweroffscreen.png | 
| 				| sleep | ${Delay} |
| power on the box | 
| 				| Click Element | powerOnButton |
| 				| sleep | ${DelayL} |
| 				| ImageHorizonLibrary.Does Exist | powerondigitalreceiverstarting.png |
| 				| sleep | ${DelayL} |
| Enter User Name and password |
|				| Input Text | input_1 | ${Username} |
|				| sleep | ${Delay} |	
|				| Input Text | input_2 | ${Password} |
| 				| Click Login |
| Click Login 	|
|				| click button | Sign In |
| scroll up 	| 
| 				| @{coordinatesScrolUp}= | Create List | ${1358} | ${100} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${coordinatesScrolUp} | 1 |
| press key two	| 
| 				| @{keytwo}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('key_two',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('key_two',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${keytwo} | 1 |
| press key One | 
| 				| @{keyone}= | Create List | ${'''+str(GlassDemoDataProvider().getCoordinates('key_one',"x"))+'''} | ${'''+str(GlassDemoDataProvider().getCoordinates('key_one',"y"))+'''} |
| 				| ImageHorizonLibrary.Click To The Above Of	| ${keyone} | 1 |
'''

    
    def writeTestCasesToFile(self):
        print "Writing test cases started"
        file= open(self.testCasefileName, 'w+')
        """this function will create test cases for each box and will write to the file"""
        file.write(self.statictext1)
        file.write ("\n")
        obj=GlassDemoDataProvider("testData/sample.xlsx")
        A6H1= obj.getvalues("A6H1")
        previousSlotID="0"  
        for SlotID in A6H1:
            testcase='''
| verifying box '''+SlotID+''' | [Documentation] | Example test |
#this test case for the box '''+SlotID+'''  
|				| Log | '''+SlotID+'''   |
|				| imagecompare.checkPreviousRecord | '''+previousSlotID +''' | '''+self.executionMode+''' |
|				| ${continueTest}= | imagecompare.returnTrue | anything |
| 				| imagecompare.resetReportValues |
| 				| imagecompare.setCurrentHEHubandSlotID | A6H1 | '''+SlotID+''' |
|				| ${SlotID} | imagecompare.getCurrentSlotID | 
| 				| imagecompare.setcontinueTest | True |
| 				| log | ${continueTest} |
| 				| sleep | ${Delay} |
| 				| Full screen Detection |
|				| Input Text | quick-select | '''+SlotID+''' |
|				| Click Element | quick-select-button |
|				| sleep | ${DelayL} |
|				| sleep | ${DelayS} |
| 				| Run Keyword And Ignore Error | scroll up |
| 				| Full screen Detection |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.ReservationVerification | Runfrompy-screenshot- | '''+SlotID+''' | '''+self.directory+'''/screenshot |
| 				| Run Keyword If | '${pass}' == 'False' | imagecompare.ReportTestStatus | '''+self.executionMode+''' |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass} |
| 				| log  | box is available |
| 				| Run Keyword If | '${pass}' == 'True' | Reserve the box |
| 				| Signal Check  |
|				| ${continueTest}= | imagecompare.getcontinueTest |
|                               | log | ${continueTest} |
| 				| Run Keyword If | '${continueTest}' == 'True' | Start Verifications |
| 				| Run Keyword If | '${continueTest}' == 'False' |  End Reservation |
| 				| imagecompare.ReportTestStatus | '''+self.executionMode+''' |
| 				| Run Keyword If | '${continueTest}' == 'False' |  Fail | msg=*HTML* ${continueTest} |'''
        
            file.write (testcase)#written test cases for all the box available in A6H1
            previousSlotID=SlotID
        ExtraCheckForLastTestcase='''
|				| imagecompare.checkPreviousRecord | '''+previousSlotID +''' | '''+self.executionMode+''' | '''
        file.write(ExtraCheckForLastTestcase)
    
        A7H2 = GlassDemoDataProvider("testData/sample.xlsx").getvalues("A7H2")
    
        previousSlotID="0" 
        for SlotID in A7H2:

            SlotID.strip(" ")
      
            testcase='''
| verifying box '''+SlotID+''' | [Documentation] | Example test |
#this test case for the box '''+SlotID+'''  
|				| Log | '''+SlotID+'''   |
|				| imagecompare.checkPreviousRecord | '''+previousSlotID +''' | '''+self.executionMode+''' |
|				| ${continueTest}= | imagecompare.returnTrue | anything |
| 				| imagecompare.resetReportValues |
| 				| imagecompare.setCurrentHEHubandSlotID | A7H2 | '''+SlotID+''' |
|				| ${SlotID} | imagecompare.getCurrentSlotID | 
| 				| imagecompare.setcontinueTest | True |
| 				| log | ${continueTest} |
| 				| sleep | ${Delay} |
| 				| Full screen Detection |
|				| Input Text | quick-select | '''+SlotID+''' |
|				| Click Element | quick-select-button |
|				| sleep | ${DelayL} |
|				| sleep | ${DelayS} |
| 				| Run Keyword And Ignore Error | scroll up |
| 				| Full screen Detection |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.ReservationVerification | Runfrompy-screenshot- | '''+SlotID+''' | '''+self.directory+'''/screenshot |
| 				| Run Keyword If | '${pass}' == 'False' | imagecompare.ReportTestStatus | '''+self.executionMode+''' |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass} |
| 				| log  | box is available|
| 				| Run Keyword If | '${pass}' == 'True' | Reserve the box |
| 				| Signal Check  |
|				| ${continueTest}= | imagecompare.getcontinueTest |
|                               | log | ${continueTest} |
| 				| Run Keyword If | '${continueTest}' == 'True' | Start Verifications |
| 				| Run Keyword If | '${continueTest}' == 'False' |  End Reservation |
| 				| imagecompare.ReportTestStatus | '''+self.executionMode+''' |
| 				| Run Keyword If | '${continueTest}' == 'False' |  Fail | msg=*HTML* ${continueTest} |'''
            if len(SlotID)>1:        
                file.write (testcase)#written test cases for all the box available in A7H2
                previousSlotID=SlotID
        ExtraCheckForLastTestcase='''
|				| imagecompare.checkPreviousRecord | '''+previousSlotID +''' | '''+self.executionMode+''' |'''
        file.write(ExtraCheckForLastTestcase)

        A7H1 = GlassDemoDataProvider("testData/sample.xlsx").getvalues("A7H1")
        file.write ("\n")
        previousSlotID="0"    
        for SlotID in A7H1:
              
            testcase='''
| verifying box '''+SlotID+''' | [Documentation] | Example test |
#this test case for the box '''+SlotID+'''  
|				| Log | '''+SlotID+'''   |
|				| imagecompare.checkPreviousRecord | '''+previousSlotID +''' | '''+self.executionMode+''' |
|				| ${continueTest}= | imagecompare.returnTrue | anything |
| 				| imagecompare.resetReportValues |
| 				| imagecompare.setCurrentHEHubandSlotID | A7H1 | '''+SlotID+''' |
|				| ${SlotID} | imagecompare.getCurrentSlotID | 
| 				| imagecompare.setcontinueTest | True |
| 				| log | ${continueTest} |
| 				| sleep | ${Delay} |
| 				| Full screen Detection |
|				| Input Text | quick-select | '''+SlotID+''' |
|				| Click Element | quick-select-button |
|				| sleep | ${DelayL} |
|				| sleep | ${DelayS} |
| 				| Run Keyword And Ignore Error | scroll up |
| 				| Full screen Detection |
| 				| Take A Screenshot |
| 				| ${pass}= | imagecompare.ReservationVerification | Runfrompy-screenshot- | '''+SlotID+''' | '''+self.directory+'''/screenshot |
| 				| Run Keyword If | '${pass}' == 'False' | imagecompare.ReportTestStatus | '''+self.executionMode+''' |
| 				| Run Keyword If | '${pass}' == 'False' | Fail | msg=*HTML* ${pass} |
| 				| log  | box is available|
| 				| Run Keyword If | '${pass}' == 'True' | Reserve the box |
| 				| Signal Check  |
|				| ${continueTest}= | imagecompare.getcontinueTest |
|                               | log | ${continueTest} |
| 				| Run Keyword If | '${continueTest}' == 'True' | Start Verifications |
| 				| Run Keyword If | '${continueTest}' == 'False' |  End Reservation |
| 				| imagecompare.ReportTestStatus | '''+self.executionMode+''' |
| 				| Run Keyword If | '${continueTest}' == 'False' |  Fail | msg=*HTML* ${continueTest} |'''
            if len(SlotID)>1:
                previousSlotID=SlotID
                file.write (testcase)
        ExtraCheckForLastTestcase='''
|				| imagecompare.checkPreviousRecord | '''+previousSlotID +''' | '''+self.executionMode+''' | '''
        file.write(ExtraCheckForLastTestcase)
                
        file.write ("\n")
        
        file.write (self.statictext2)
        file.write ("\n")
        file.close()

    def ExecuteScript(self):
        print "starting execution"
        """This function will start execution of the script. No input required """
        call_command= ['C:/Python27/python.exe', '-m', 'robot', 'C:\\Python27\\Scripts\\'+self.directory+'\\Runfrompy.txt']
        call(call_command)
    def getDatafromOldreport(self):
        pass
        
    def resetReport(self):
        
        self.reportPath=GlassDemoDataProvider().getFilepaths('Report_src')
        if os.path.isfile(self.reportPath) and os.access(self.reportPath, os.R_OK):
            xfile = openpyxl.load_workbook(self.reportPath)
            sheet = xfile.get_sheet_by_name('Master Data')
            for i in range(2,130):
                sheet[str("A")+str(i)]=""
                sheet[str("B")+str(i)]=""
                sheet[str("C")+str(i)]=""
                sheet[str("D")+str(i)]=""
            Summarysheet = xfile.get_sheet_by_name('Summary')
            Summarysheet[str("J")+str("2")]=self.Date
            print 'reset report called date value is ',self.Date
            
                
                
            
            xfile.save(self.reportPath)
            print "report file reset done"
        else: print"report file not found"
    def editReportFile(self,sheet,column,row,data):
        """Send data in the format sheet,cloumn,row,data ex: ("Master Data","C","3","PASS") """
        if os.path.isfile(self.reportPath) and os.access(self.reportPath, os.R_OK):
            xfile = openpyxl.load_workbook(self.reportPath)
            sheet = xfile.get_sheet_by_name(sheet)
            sheet[str(column)+str(int(row))]=data
            xfile.save(self.reportPath)        
    def readDatafromExcel(self,sheet_name,column,row):
        self.reportPath=GlassDemoDataProvider().getFilepaths('Report_src')
        if os.path.isfile(self.reportPath) and os.access(self.reportPath, os.R_OK):
            xfile = openpyxl.load_workbook(self.reportPath,data_only=True)
            sheet = xfile[sheet_name]
            return (sheet[str(column)+str(int(row))].value)
        else:print "report file not found"
        
    def getDatafromOldreport(self):
        current_raw=int(self.readDatafromExcel('Summary',"z","777"))#4
        oldExecutionDate=str(self.readDatafromExcel('Summary',"J","2"))
        print "oldexecutiondate",oldExecutionDate
        


        
        oldpasscount= str(self.readDatafromExcel('Summary',"K","6"))
        
        print "oldpasscount",oldpasscount, type(oldpasscount)
        
        
        oldStbIssueCount=str(self.readDatafromExcel('Summary',"L","6"))
        print "oldStbIssueCount",oldStbIssueCount, type(oldStbIssueCount)
        
        
        oldGlasIssueCount=str(self.readDatafromExcel('Summary',"M","6"))
        print "oldGlasIssueCount",oldGlasIssueCount, type(oldGlasIssueCount)
        
        
        oldSkippedIssueCount=str(self.readDatafromExcel('Summary',"N","6"))
        print "oldSkippedIssueCount",oldSkippedIssueCount, type(oldSkippedIssueCount)
        
        
        oldTotalIssueCount=str(self.readDatafromExcel('Summary',"O","6"))
        print "oldTotalIssueCount",oldTotalIssueCount, type(oldTotalIssueCount)


        self.editReportFile('Summary',"B",current_raw,oldExecutionDate)
        self.editReportFile('Summary',"C",current_raw,str(oldpasscount))
        self.editReportFile('Summary',"D",current_raw,oldStbIssueCount)
        self.editReportFile('Summary',"E",current_raw,oldGlasIssueCount)
        self.editReportFile('Summary',"F",current_raw,oldSkippedIssueCount)
        
        self.editReportFile('Summary',"G",current_raw,oldTotalIssueCount)
        self.editReportFile('Summary',"Z","777",str((current_raw+1)))
        
        
           
    def copyReportFiles(self):
        """this function will preserve repot ang log file from overriting the in next run. No input required """ 
        src_of_log='C:/Python27/Scripts/log.html'
        src_of_report='C:/Python27/Scripts/report.html'
        dst_of_log='C:/Python27/Scripts/'+self.directory+'/log.html'
        dst_of_report='C:/Python27/Scripts/'+self.directory+'/report.html'
        copyfile(src_of_log, dst_of_log)
        copyfile(src_of_report, dst_of_report)
        src_of_excel_report='C:/Python27/Scripts/testData/Glas Automation Status.xlsx'
        dst_of_excel_report='C:/Python27/Scripts/'+self.directory+'/Glas Automation Status'+self.directory+'.xlsx'
        copyfile(src_of_excel_report, dst_of_excel_report)

        

    def makeEmailBodyDataForJenkins(self):

        print "makeEmailBodyDataForJenkins() called"
        emailBody = open("testData/emailbody.txt",'w+')
        emailBody.write("Date and time of execution 	:: ")
        emailBody.write(self.directory)
        emailBody.write("\n")
        emailBody.write("Total test executed 		:: ")
        count= self.readDatafromExcel("Summary","O","6")
        if type(count)==type(None):
            count=0
        emailBody.write(str(count))
        print count
        emailBody.write("\n")
        emailBody.write("Total test passed 		:: ")
        count= self.readDatafromExcel("Summary","K","6")
        if type(count)==type(None):
            count=0
        emailBody.write(str(count))
        print count
        emailBody.write("\n")
        emailBody.write("Total test Failed 		:: ")
        count= self.readDatafromExcel("Summary","L","6")
        #print"type of none is",type(None)
        if type(count) == type(None):
            count= 0
        else:
            count=int(count)
        
        count1=(self.readDatafromExcel("Summary","M","6"))

        if type(count1) == type(None):
            count1=0
        else:
            count1=int(count1)
        count= count+count1
        print count
        emailBody.write(str(count))
        emailBody.write("\n")
        emailBody.write("Total test Skipped 		:: ")
        count= self.readDatafromExcel("Summary","N","6")
        if type(count) == type(None):
            print "issue", count
            count=0
        print count    
        emailBody.write(str(count))
        emailBody.write("\n")
        emailBody.close()
    def makeEmailBodyDataForJenkinsFromReportOverviewFile(self):
        with open("testData/reportOverView.txt", "r") as ins:
            array = ''
            for line in ins:
                array=array+line
        countPass=array.count('PASS')
        countSkipped=array.count('SKIPPED')
        countSTBissue=array.count('STB_ISSUE')
        countGlasissue=array.count('GLAS_ISSUE')
        total=countPass+countSkipped+countSTBissue+countGlasissue
        print "count is",countPass,countSkipped,countSTBissue,countGlasissue,total
        print "makeEmailBodyDataForJenkins() called"
        emailBody = open("testData/emailbody.txt",'w+')
        emailBody.write("Date and time of execution 	:: ")
        emailBody.write(self.directory)
        emailBody.write("\n")
        emailBody.write("Total test executed 		:: ")
        emailBody.write(str(total))
        emailBody.write("\n")
        emailBody.write("Total test passed 		:: ")
        emailBody.write(str(countPass))

        emailBody.write("\n")
        emailBody.write("Total test Failed 		:: ")
        failed=countSTBissue+countGlasissue
        emailBody.write(str(failed))
        emailBody.write("\n")
        emailBody.write("Total test Skipped 		:: ")
    
        emailBody.write(str(countSkipped))
        emailBody.write("\n")
        emailBody.close()
        
    def clearEmailBodyData(self):
        emailBody = open("testData/emailbody.txt",'w+')
        emailBody.close()
        
        
        
    def main(self):
        """This is the start point of execution, this functon will detemine the flow. No input required """
        self.createTimestamp()
        if (self.executionMode=='jenkins'):
            self.clearEmailBodyData()
        self.ceateExecutionDirectory()
        self.createScreenshotDirectory()
        self.createTestCaseFile()
        self.createReservedBoxes()
        self.createListOfGuideFailedBoxes()
        self.createReportOverviewFile()
        self.getScreenShotPrefix()
        self.setStaticText()
        self.writeTestCasesToFile()
        if (self.executionMode=='batch_run') or (self.executionMode=='jenkins'):

            #self.getDatafromOldreport()
            self.resetReport()
            pass
            

        self.ExecuteScript()
        self.copyReportFiles()
        if (self.executionMode=='jenkins'):
            self.makeEmailBodyDataForJenkinsFromReportOverviewFile()
            
            pass

        print "created"
    def __init__(self):
        self.main()

#calling main()        
RunRobotFromPython()

